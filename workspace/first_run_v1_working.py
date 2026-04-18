#!/usr/bin/env python3
"""
Pokemon Booster Box Tracker — Monthly updater
"""

import re
import datetime
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from curl_cffi import requests as cffi_requests

SPREADSHEET = "/root/.openclaw/workspace/pokemon-tracker.xlsx"
DATA_ROWS = range(2, 45)
WIZARD_BASE = "https://www.pokemonwizard.com"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

DAWNGLARE_EXACT_MAP = {
    "ascended hereos":             "ascended heroes",
    "celebrations 25th":           "celebrations",
    "champions path":              "champion's path",
    "s&v base set":                "scarlet & violet",
    "mega evolution (enhanced)":   "mega evolution enhanced",
    "journey together (enhanced)": "journey together enhanced",
    "evolutions":                  "xy evolutions",
}

ETB_ONLY_SETS = {
    "prismatic evolutions": "prismatic evolutions elite trainer box",
    "black bolt":           "black bolt elite trainer box",
    "white flare":          "white flare elite trainer box",
    "ascended hereos":      "ascended heroes elite trainer box",
    "celebrations 25th":    "celebrations elite trainer box",
    "champions path":       "champion's path elite trainer box",
    "shining fates":        "shining fates elite trainer box",
    "crown zenith":         "crown zenith elite trainer box",
    "paldean fates":        "paldean fates elite trainer box",
    "151":                  "151 elite trainer box",
    "hidden fates":         "hidden fates elite trainer box",
    "shrouded fable":       "shrouded fable elite trainer box",
}

# Spreadsheet name -> Wizard set name (only where they differ)
WIZARD_NAME_MAP = {
    "s&v base set":                "Scarlet & Violet Base Set",
    "151":                         "Scarlet & Violet 151",
    "ascended hereos":             "Ascended Heroes",
    "celebrations 25th":           "Celebrations",
    "champions path":              "Champion's Path",
    "sword and shield":            "Sword & Shield",
    "mega evolution (enhanced)":   "Mega Evolution",
    "journey together (enhanced)": "Journey Together",
}

def pct_fill(pct):
    if pct >= 1.20: return PatternFill("solid", fgColor="00B050")
    if pct >= 1.05: return PatternFill("solid", fgColor="92D050")
    if pct >= 0.90: return PatternFill("solid", fgColor="FFFF00")
    if pct >= 0.75: return PatternFill("solid", fgColor="FFC000")
    return PatternFill("solid", fgColor="FF0000")

def get_usd_to_gbp():
    try:
        r = requests.get("https://open.er-api.com/v6/latest/USD", timeout=10)
        return r.json()["rates"]["GBP"]
    except Exception as e:
        print(f"  [WARN] Exchange rate fetch failed: {e} — using 0.79")
        return 0.79

# ── Dawnglare ─────────────────────────────────────────────────────────────────
def fetch_dawnglare_prices():
    print("  Fetching dawnglare.com ...")
    try:
        resp = cffi_requests.get("https://pokemon.dawnglare.com/?p=boxprice", impersonate="chrome110", timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [ERROR] dawnglare fetch failed: {e}")
        return {}

    soup = BeautifulSoup(resp.text, "html.parser")
    prices = {}
    for a_tag in soup.find_all("a", href=True):
        name = a_tag.get_text(strip=True)
        td = a_tag.find_parent("td")
        if not td:
            continue
        next_td = td.find_next_sibling("td")
        if not next_td:
            continue
        span = next_td.find("span", class_=re.compile(r'^pi\d+$'))
        if not span:
            continue
        price_text = span.get_text(strip=True).replace("$", "").replace(",", "")
        try:
            prices[name.lower()] = float(price_text)
        except ValueError:
            continue

    print(f"  dawnglare: found {len(prices)} entries")
    return prices

def find_booster_box_price(sheet_name, dawnglare_prices):
    key = sheet_name.strip().lower()
    mapped = DAWNGLARE_EXACT_MAP.get(key, key)

    if mapped + " booster box" in dawnglare_prices:
        return dawnglare_prices[mapped + " booster box"], "BB"
    if mapped + " enhanced booster box" in dawnglare_prices:
        return dawnglare_prices[mapped + " enhanced booster box"], "Enhanced BB"

    key_words = set(mapped.split()) - {"the", "and", "&", "of", "a"}
    bb_candidates = {n: p for n, p in dawnglare_prices.items() if "booster box" in n}
    best_score, best_name, best_price = 0, None, None
    for dg_name, price in bb_candidates.items():
        score = len(key_words & set(dg_name.split()))
        if score > best_score:
            best_score, best_name, best_price = score, dg_name, price
    if best_score >= 2:
        return best_price, f"BB~{best_name}"

    if key in ETB_ONLY_SETS:
        etb_fragment = ETB_ONLY_SETS[key]
        for dg_name, price in dawnglare_prices.items():
            if etb_fragment in dg_name and "exclusive" not in dg_name and "set of" not in dg_name:
                return price * 4, "ETBx4"

    return None, "NOT FOUND"

# ── PokemonWizard ─────────────────────────────────────────────────────────────
def fetch_wizard_set_index():
    """Scrape /sets page and return dict: set_name_lower -> full URL"""
    print("  Fetching PokemonWizard set index ...")
    try:
        r = requests.get(f"{WIZARD_BASE}/sets", timeout=15, headers=HEADERS)
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"  [ERROR] Wizard index fetch failed: {e}")
        return {}

    index = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not re.match(r'^/sets/\d+/', href):
            continue
        name = a.get_text(strip=True)
        if name and name != "Buy":
            index[name.lower()] = WIZARD_BASE + href
    print(f"  Wizard index: {len(index)} sets")
    return index

def find_wizard_url(sheet_name, wizard_index):
    """Match spreadsheet set name to a Wizard URL."""
    key = sheet_name.strip().lower()
    # Check explicit map first
    mapped = WIZARD_NAME_MAP.get(key, sheet_name).lower()

    if mapped in wizard_index:
        return wizard_index[mapped]
    if key in wizard_index:
        return wizard_index[key]

    # Fuzzy: find best word-overlap match
    key_words = set(mapped.split()) - {"the", "and", "&", "of", "a"}
    best_score, best_url = 0, None
    for wname, url in wizard_index.items():
        score = len(key_words & set(wname.split()))
        if score > best_score:
            best_score, best_url = score, url
    if best_score >= 2:
        return best_url
    return None

def fetch_wizard_data(set_url, rate):
    """Fetch set page, return (total_value_gbp, top3_names_str)."""
    try:
        r = requests.get(set_url, timeout=20, headers=HEADERS)
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"    [WARN] Wizard page fetch failed: {e}")
        return None, None

    # Total value — find "Total Value" row, grab text-danger or text-success span
    total_val = None
    for strong in soup.find_all("strong"):
        if strong.get_text(strip=True) == "Total Value":
            row = strong.find_parent("tr")
            if row:
                span = row.find("span", class_=re.compile(r"text-(danger|success)"))
                if span:
                    val_text = span.get_text(strip=True).replace("$", "").replace(",", "").split()[0]
                    try:
                        total_val = round(float(val_text) * rate, 2)
                    except ValueError:
                        pass
            break

    # Top 3 chase cards — parse all card rows, sort by price, take top 3
    cards = []
    for row in soup.select("tr"):
        tds = row.find_all("td")
        if len(tds) < 5:
            continue
        price_text = tds[4].get_text(strip=True).replace("$", "").replace(",", "")
        try:
            price = float(price_text)
        except ValueError:
            continue
        raw_name = tds[1].get_text(strip=True)
        name = re.sub(r'\s+\d+\s+\d+\s*$', '', raw_name).strip()
        if name:
            cards.append((name, price))

    cards.sort(key=lambda x: x[1], reverse=True)
    top3_str = ", ".join(n for n, _ in cards[:3]) if cards else None
    return total_val, top3_str

# ── Sheet helpers ─────────────────────────────────────────────────────────────
def new_sheet_name(dt=None):
    return (dt or date.today()).strftime("%b %y")

def prev_sheet_name(dt=None):
    dt = dt or date.today()
    prev = dt.replace(day=1) - datetime.timedelta(days=1)
    return prev.strftime("%b %y")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    today = date.today()
    new_name = new_sheet_name(today)
    prev_name = prev_sheet_name(today)

    print(f"=== Pokemon Tracker === {today} ===")
    print(f"  New sheet : {new_name}")
    print(f"  Prev sheet: {prev_name}")

    wb = load_workbook(SPREADSHEET)

    if new_name in wb.sheetnames:
        print(f"  Sheet '{new_name}' already exists — updating in place.")
        ws = wb[new_name]
    else:
        src = wb[prev_name] if prev_name in wb.sheetnames else wb.worksheets[0]
        if prev_name not in wb.sheetnames:
            print(f"  [WARN] '{prev_name}' not found; copying from '{src.title}'")
        ws = wb.copy_worksheet(src)
        ws.title = new_name
        wb.move_sheet(ws, offset=len(wb.sheetnames))
        print(f"  Created sheet '{new_name}' from '{src.title}'")

    rate = get_usd_to_gbp()
    print(f"  USD->GBP rate: {rate:.4f}")

    dawnglare = fetch_dawnglare_prices()
    wizard_index = fetch_wizard_set_index()

    print("\n  Updating rows ...")
    found_d, found_ef = 0, 0

    for row in DATA_ROWS:
        set_name = ws.cell(row=row, column=3).value
        if not set_name:
            continue
        set_name = str(set_name).strip()
        print(f"\n  Row {row}: {set_name}")

        # Column D — booster box price
        usd_price, note = find_booster_box_price(set_name, dawnglare)
        if usd_price is not None:
            gbp = round(usd_price * rate, 2)
            ws.cell(row=row, column=4).value = gbp
            print(f"    D: £{gbp:.2f}  [{note}]")
            found_d += 1
        else:
            print(f"    D: NOT FOUND")

        # Columns E & F — set value and chase cards
        wizard_url = find_wizard_url(set_name, wizard_index)
        if wizard_url:
            set_value, chase = fetch_wizard_data(wizard_url, rate)
            if set_value is not None:
                ws.cell(row=row, column=5).value = set_value
                print(f"    E: £{set_value:.2f}")
                found_ef += 1
            else:
                print(f"    E: value not found on Wizard page")
            if chase:
                ws.cell(row=row, column=6).value = chase
                print(f"    F: {chase}")
        else:
            print(f"    E/F: no Wizard match")

        # Formulas
        ws.cell(row=row, column=7).value = f'=IF(D{row}>0,E{row}/D{row},"")'
        ws.cell(row=row, column=9).value = f'=IF(D{row}>0,E{row}/D{row},"")'

        # Colour code G
        if usd_price and set_value:
            bb_gbp = usd_price * rate
            if bb_gbp > 0:
                try:
                    ws.cell(row=row, column=7).fill = pct_fill(set_value / bb_gbp)
                except (TypeError, ZeroDivisionError):
                    pass

    wb.save(SPREADSHEET)
    # Copy to jaytee's home so it can be downloaded via scp
    import shutil
    shutil.copy2(SPREADSHEET, "/home/jaytee/pokemon-tracker.xlsx")
    import os
    os.chmod("/home/jaytee/pokemon-tracker.xlsx", 0o644)
    print(f"\n{'='*50}")
    print(f"Done! Saved: {SPREADSHEET}")
    print(f"  Copied to: /home/jaytee/pokemon-tracker.xlsx")
    print(f"  Column D populated: {found_d}/43")
    print(f"  Columns E&F populated: {found_ef}/43")

if __name__ == "__main__":
    main()
